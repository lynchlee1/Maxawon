from __future__ import annotations

import csv
import subprocess
import sys
from pathlib import Path
from tkinter import BOTH, DISABLED, END, NORMAL, StringVar, Tk, filedialog, messagebox, ttk

from cretop_data_reader.scrapling_adapter import check_scrapling


CRETOP_URL = "https://www.cretop.com/"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
PROFILE_DIR = PROJECT_ROOT / ".chrome-profile"


def find_chrome() -> str | None:
    candidates = [
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
        Path.home() / r"AppData\Local\Google\Chrome\Application\chrome.exe",
    ]

    for candidate in candidates:
        if candidate.exists():
            return str(candidate)

    return None


def read_excel_preview(path: Path, limit: int = 20) -> tuple[list[str], list[list[str]]]:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        last_error: UnicodeDecodeError | None = None
        for encoding in ("utf-8-sig", "cp949"):
            try:
                with path.open("r", encoding=encoding, newline="") as file:
                    reader = csv.reader(file)
                    rows = list(reader)
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            raise RuntimeError("CSV 파일 인코딩을 읽지 못했습니다.") from last_error
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "엑셀(.xlsx) 파일을 읽으려면 openpyxl이 필요합니다. "
                "터미널에서 `pip install -r requirements.txt`를 실행하세요."
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [
            ["" if value is None else str(value) for value in row]
            for row in sheet.iter_rows(max_row=limit + 1, values_only=True)
        ]
        workbook.close()

    if not rows:
        return [], []

    headers = rows[0]
    body = rows[1 : limit + 1]
    return headers, body


class CretopDataReaderApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.root.title("Cretop Data Reader")
        self.root.geometry("920x620")
        self.root.minsize(760, 520)

        self.excel_path: Path | None = None
        self.login_status = StringVar(value="로그인 전")
        self.file_status = StringVar(value="엑셀 파일 미선택")
        self.progress_status = StringVar(value="대기 중")
        self.scrapling_status = StringVar(value="확인 전")

        self._build()

    def _build(self) -> None:
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(2, weight=1)

        top = ttk.Frame(self.root, padding=16)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        ttk.Label(top, text="Cretop Data Reader", font=("", 16, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w"
        )
        ttk.Label(top, text="수동 로그인 후 엑셀 검색 대상을 처리하는 업무용 도구").grid(
            row=1, column=0, columnspan=3, sticky="w", pady=(4, 0)
        )

        actions = ttk.Frame(self.root, padding=(16, 0, 16, 12))
        actions.grid(row=1, column=0, sticky="ew")
        actions.columnconfigure(3, weight=1)

        ttk.Button(actions, text="Chrome 열기", command=self.open_chrome).grid(
            row=0, column=0, padx=(0, 8), sticky="w"
        )
        ttk.Button(actions, text="로그인 완료", command=self.mark_login_done).grid(
            row=0, column=1, padx=(0, 8), sticky="w"
        )
        ttk.Button(actions, text="엑셀 선택", command=self.pick_excel).grid(
            row=0, column=2, padx=(0, 8), sticky="w"
        )
        ttk.Button(actions, text="Scrapling 확인", command=self.check_scrapling_status).grid(
            row=0, column=3, padx=(0, 8), sticky="w"
        )
        self.start_button = ttk.Button(actions, text="처리 시작", command=self.start_processing)
        self.start_button.grid(row=0, column=4, sticky="e")

        status = ttk.Frame(self.root, padding=(16, 0, 16, 12))
        status.grid(row=2, column=0, sticky="nsew")
        status.columnconfigure(0, weight=1)
        status.rowconfigure(1, weight=1)

        status_grid = ttk.Frame(status)
        status_grid.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        status_grid.columnconfigure(1, weight=1)

        ttk.Label(status_grid, text="로그인 상태").grid(row=0, column=0, sticky="w", padx=(0, 10))
        ttk.Label(status_grid, textvariable=self.login_status).grid(row=0, column=1, sticky="w")
        ttk.Label(status_grid, text="파일 상태").grid(row=1, column=0, sticky="w", padx=(0, 10))
        ttk.Label(status_grid, textvariable=self.file_status).grid(row=1, column=1, sticky="w")
        ttk.Label(status_grid, text="진행 상태").grid(row=2, column=0, sticky="w", padx=(0, 10))
        ttk.Label(status_grid, textvariable=self.progress_status).grid(row=2, column=1, sticky="w")
        ttk.Label(status_grid, text="Scrapling").grid(row=3, column=0, sticky="w", padx=(0, 10))
        ttk.Label(status_grid, textvariable=self.scrapling_status).grid(row=3, column=1, sticky="w")

        preview_frame = ttk.LabelFrame(status, text="엑셀 미리보기", padding=10)
        preview_frame.grid(row=1, column=0, sticky="nsew")
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)

        self.preview = ttk.Treeview(preview_frame, show="headings", height=12)
        self.preview.grid(row=0, column=0, sticky="nsew")

        y_scroll = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(preview_frame, orient="horizontal", command=self.preview.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.preview.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        log_frame = ttk.LabelFrame(self.root, text="로그", padding=10)
        log_frame.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 16))
        log_frame.columnconfigure(0, weight=1)

        self.log = ttk.Treeview(log_frame, columns=("message",), show="headings", height=6)
        self.log.heading("message", text="메시지")
        self.log.column("message", width=800, stretch=True)
        self.log.grid(row=0, column=0, sticky="ew")

        self._set_start_enabled()
        self.add_log("Chrome을 열고 직접 로그인한 뒤 '로그인 완료'를 누르세요.")
        self.check_scrapling_status(show_popup=False)

    def open_chrome(self) -> None:
        chrome = find_chrome()
        if chrome is None:
            messagebox.showerror("Chrome 없음", "이 PC에서 Chrome 실행 파일을 찾지 못했습니다.")
            return

        PROFILE_DIR.mkdir(exist_ok=True)
        subprocess.Popen(
            [
                chrome,
                f"--user-data-dir={PROFILE_DIR}",
                "--new-window",
                CRETOP_URL,
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        self.progress_status.set("Chrome 실행됨")
        self.add_log("Chrome을 열었습니다. Cretop에서 직접 로그인하세요.")

    def mark_login_done(self) -> None:
        self.login_status.set("로그인 완료")
        self.progress_status.set("엑셀 파일 선택 대기")
        self.add_log("사용자가 로그인 완료를 확인했습니다.")
        self._set_start_enabled()

    def check_scrapling_status(self, show_popup: bool = True) -> None:
        status = check_scrapling()
        self.scrapling_status.set(status.message)
        self.add_log(status.message)

        if show_popup:
            if status.installed:
                messagebox.showinfo("Scrapling 확인", status.message)
            else:
                messagebox.showwarning("Scrapling 필요", status.message)

    def pick_excel(self) -> None:
        selected = filedialog.askopenfilename(
            title="검색 대상 파일 선택",
            filetypes=[
                ("Excel or CSV", "*.xlsx *.xlsm *.csv"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if not selected:
            return

        path = Path(selected)
        try:
            headers, rows = read_excel_preview(path)
        except Exception as exc:
            messagebox.showerror("파일 읽기 실패", str(exc))
            return

        self.excel_path = path
        self.file_status.set(str(path))
        self.progress_status.set("엑셀 파일 로드됨")
        self.render_preview(headers, rows)
        self.add_log(f"검색 대상 파일을 불러왔습니다: {path.name}")
        self._set_start_enabled()

    def render_preview(self, headers: list[str], rows: list[list[str]]) -> None:
        self.preview.delete(*self.preview.get_children())
        column_ids = [f"col_{index}" for index in range(len(headers))]
        self.preview["columns"] = column_ids

        for index, column_id in enumerate(column_ids):
            header = headers[index] or f"Column {index + 1}"
            self.preview.heading(column_id, text=header)
            self.preview.column(column_id, width=140, minwidth=80, stretch=True)

        for row in rows:
            values = row + [""] * (len(headers) - len(row))
            self.preview.insert("", END, values=values[: len(headers)])

    def start_processing(self) -> None:
        if self.login_status.get() != "로그인 완료":
            messagebox.showwarning("로그인 필요", "먼저 Cretop에 로그인한 뒤 '로그인 완료'를 누르세요.")
            return
        if self.excel_path is None:
            messagebox.showwarning("파일 필요", "검색 대상 엑셀 파일을 선택하세요.")
            return
        scrapling = check_scrapling()
        if not scrapling.installed:
            self.scrapling_status.set(scrapling.message)
            messagebox.showwarning("Scrapling 필요", scrapling.message)
            return

        self.progress_status.set("구현 대기")
        self.add_log("Scrapling 기반 검색 처리는 아직 구현되지 않았습니다. 중복 후보 처리 규칙을 먼저 확정해야 합니다.")
        messagebox.showinfo(
            "아직 구현 전",
            "Scrapling 기반 엑셀 검색 처리는 아직 보류 상태입니다.\n"
            "기업명/법인번호 컬럼, 출력 항목, 중복 후보 선택 기준을 확정한 뒤 구현하세요.",
        )

    def add_log(self, message: str) -> None:
        self.log.insert("", END, values=(message,))
        children = self.log.get_children()
        if children:
            self.log.see(children[-1])

    def _set_start_enabled(self) -> None:
        if self.login_status.get() == "로그인 완료" and self.excel_path is not None:
            self.start_button.configure(state=NORMAL)
        else:
            self.start_button.configure(state=DISABLED)


def main() -> int:
    root = Tk()
    CretopDataReaderApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
