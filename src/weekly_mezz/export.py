import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from weekly_mezz.dart import fetch_previous_family_rcept_no
from weekly_mezz.parser import parse_report_document

INCLUSION_KEYWORDS = [["전환사채", "교환사채", "신주인수권부사채"], "발행"]
DART_MAIN_URL = "https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}"
KIND_DISCLOSURE_VIEWER_URL = "https://kind.krx.co.kr/common/disclsviewer.do"
# Developer-only diagnostic output. Keep this False for release builds.
ENABLE_AUDIT_JSON = False
DEFAULT_PARSE_WORKERS = 4

HEADER_FILL = PatternFill(fill_type="solid", fgColor="00D084")
HEADER_FONT = Font(bold=True, color="06110B", size=9)
BODY_FONT = Font(size=9)
BORDER_SIDE = Side(style="thin", color="263244")
HEADER_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
BODY_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BLANK_CELL_VALUE = "-"

SHARE_TYPE_SUFFIX_PATTERN = re.compile(
    r"\s*(?:(?:기명식|무기명식)\s*)?(?:보통주식?|우선주식?|전환우선주식?|상환우선주식?|상환전환우선주식?|종류주식)\s*$"
)
CORP_DESIGNATOR_PATTERN = re.compile(r"\(주\)|㈜|주식회사")
MATCH_NORMALIZE_PATTERN = re.compile(r"[^0-9A-Za-z가-힣]")
TRUSTEE_NAME_PATTERN = re.compile(r"신탁|수탁")

ISSUER_MARKET_LABELS = {
    "Y": "코스피",
    "K": "코스닥",
    "N": "코넥스",
    "E": "기타",
}

COLUMN_SPECS = [
    {"header": "헤더", "key": "report_header", "width": 12, "align": "center"},
    {"header": "최초공시일", "key": "initial_filing_date_display", "width": 12, "align": "center"},
    {"header": "공시일", "key": "filing_date_display", "width": 12, "align": "center"},
    {"header": "납입일", "key": "issue_date_display", "width": 12, "align": "center"},
    {"header": "발행사 기업명", "key": "issuer_company_name", "width": 20, "align": "left", "margin": True},
    {"header": "상장시장", "key": "issuer_market", "width": 10, "align": "center"},
    {"header": "교환대상 기업명", "key": "target_company_name", "width": 20, "align": "left", "margin": True},
    {"header": "종류", "key": "security_type", "width": 7, "align": "center"},
    {"header": "벤처여부", "key": "venture_blank", "width": 10, "align": "center"},
    {"header": "시가총액", "key": "market_cap_eok", "width": 12, "align": "right", "margin": True},
    {"header": "발행금액", "key": "issue_amount_eok", "width": 12, "align": "right", "margin": True},
    {"header": "행사가액", "key": "strike_price", "width": 12, "align": "right", "margin": True},
    {"header": "할증률", "key": "premium_text", "width": 34, "align": "left", "margin": True},
    {"header": "만기", "key": "maturity_term_text", "width": 9, "align": "center"},
    {"header": "PUT", "key": "put_text", "width": 8, "align": "left"},
    {"header": "표면이자율", "key": "coupon_rate_text", "width": 11, "align": "right"},
    {"header": "만기이자율", "key": "maturity_rate_text", "width": 12, "align": "left"},
    {"header": "CALL", "key": "call_blank", "width": 8, "align": "center"},
    {"header": "Refixing", "key": "refixing_floor_text", "width": 12, "align": "right", "margin": True},
    {"header": "리픽싱사유", "key": "refixing_reason", "width": 34, "align": "left", "margin": True},
    {"header": "투자자", "key": "investors_text", "width": 34, "align": "left", "margin": True},
    {"header": "섹터", "key": "sector_blank", "width": 12, "align": "left", "margin": True},
    {"header": "당사검토", "key": "internal_review", "width": 12, "align": "left", "margin": True},
    {"header": "주관", "key": "lead_blank", "width": 12, "align": "left", "margin": True},
    {"header": "URL", "key": "disclosure_url", "width": 52, "align": "left", "margin": True},
]

NUMERIC_KEYS = {
    "market_cap_eok",
    "issue_amount_eok",
    "strike_price",
    "coupon_rate_pct",
    "maturity_rate_pct",
    "put_term_years",
    "put_ytp_pct",
    "call_term_years",
    "call_ytc_pct",
    "call_cap_pct",
    "refixing_cycle_months",
    "refixing_price_won",
    "refixing_floor_pct",
}
INTEGER_FORMAT_KEYS = {"market_cap_eok", "issue_amount_eok", "strike_price"}
DECIMAL_FORMAT_KEYS = NUMERIC_KEYS - INTEGER_FORMAT_KEYS


@dataclass
class ExportResult:
    output_path: Path
    audit_path: Path | None
    raw_path: Path | None
    summary: dict


def default_output_path(filename: str = "mezzanine_reports.xlsx") -> Path:
    desktop = Path.home() / "Desktop"
    base_dir = desktop if desktop.exists() else Path.home()
    return base_dir / filename


def ensure_parent_dir(path) -> Path:
    value = Path(path).expanduser()
    value.parent.mkdir(parents=True, exist_ok=True)
    return value


def should_include_report(report: dict) -> bool:
    if report.get("corp_cls", "") not in {"Y", "K"}:
        return False
    report_nm = report.get("report_nm", "")
    for keyword_group in INCLUSION_KEYWORDS:
        if isinstance(keyword_group, list):
            if not any(keyword in report_nm for keyword in keyword_group):
                return False
        elif keyword_group not in report_nm:
            return False
    return True


def filter_reports(reports: list[dict]) -> list[dict]:
    return [report for report in reports if should_include_report(report)]


def format_issuer_stock_code(stock_code) -> str:
    value = normalize_stock_code(stock_code)
    return f"A{value}" if value else ""


def normalize_stock_code(stock_code) -> str:
    value = (stock_code or "").strip().upper()
    return value[1:] if value.startswith("A") else value


def normalize_stock_code_key(stock_code) -> str:
    value = normalize_stock_code(stock_code)
    return value.zfill(6) if value.isdigit() else value


def clean_target_stock_name(value) -> str:
    text = (value or "").strip()
    if not text or text == "-":
        return ""
    text = SHARE_TYPE_SUFFIX_PATTERN.sub("", text)
    text = CORP_DESIGNATOR_PATTERN.sub("", text)
    text = re.sub(r"\s+", " ", text).strip(" ,")
    normalized = MATCH_NORMALIZE_PATTERN.sub("", text).upper()
    if normalized in {"발행회사", "발행회사의"}:
        return ""
    return text


def normalize_company_name_for_match(value) -> str:
    cleaned = clean_target_stock_name(value)
    return MATCH_NORMALIZE_PATTERN.sub("", cleaned).upper() if cleaned else ""


def build_company_stock_code_map(corp_code_entries: list[dict]) -> dict:
    mapping = {}
    for entry in corp_code_entries or []:
        stock_code = (entry.get("stock_code") or "").strip()
        corp_name = (entry.get("corp_name") or "").strip()
        if not stock_code or not corp_name:
            continue
        normalized = normalize_company_name_for_match(corp_name)
        if normalized and normalized not in mapping:
            mapping[normalized] = stock_code
    return mapping


def parse_market_cap_value(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, str):
        value = value.strip().replace(",", "")
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def fetch_market_cap_map(progress_callback=None) -> dict:
    try:
        import FinanceDataReader as fdr
    except ImportError as exc:
        if progress_callback:
            progress_callback(f"FinanceDataReader import failed; market caps will stay blank: {exc}")
        return {}

    try:
        listing = fdr.StockListing("KRX")
    except Exception as exc:
        if progress_callback:
            progress_callback(f"FinanceDataReader KRX listing failed; market caps will stay blank: {exc}")
        return {}

    code_column = next((column for column in ("Code", "Symbol", "종목코드") if column in listing.columns), "")
    market_cap_column = next((column for column in ("Marcap", "MarketCap", "시가총액") if column in listing.columns), "")
    if not code_column or not market_cap_column:
        if progress_callback:
            progress_callback("FinanceDataReader KRX listing did not include Code/Marcap columns; market caps will stay blank")
        return {}

    market_caps = {}
    for _, item in listing.iterrows():
        code = normalize_stock_code_key(item.get(code_column))
        market_cap_won = parse_market_cap_value(item.get(market_cap_column))
        if code and market_cap_won is not None:
            market_caps[code] = round(market_cap_won / 100_000_000)
    return market_caps


def infer_security_type(report_nm, parsed_type) -> str:
    if parsed_type and parsed_type != "N/A":
        return parsed_type
    report_nm = report_nm or ""
    if "전환사채" in report_nm:
        return "CB"
    if "교환사채" in report_nm:
        return "EB"
    if "신주인수권부사채" in report_nm:
        return "BW"
    return ""


def build_dart_link(rcept_no) -> str:
    value = (rcept_no or "").strip()
    return DART_MAIN_URL.format(rcept_no=value) if value else ""


def build_kind_link(acpt_no, doc_no="") -> str:
    value = (acpt_no or "").strip()
    if not value:
        return ""
    return f"{KIND_DISCLOSURE_VIEWER_URL}?{urlencode({'method': 'search', 'acptno': value, 'docno': doc_no or '', 'viewerhost': '', 'viewerport': ''})}"


def build_disclosure_link(report: dict) -> str:
    acpt_no = report.get("acpt_no")
    if acpt_no:
        return build_kind_link(acpt_no, report.get("doc_no") or "")
    return build_dart_link(report.get("rcept_no"))


def extract_filing_date_from_rcept_no(rcept_no) -> str:
    value = (rcept_no or "").strip()
    if not re.fullmatch(r"\d{14}", value):
        return ""
    try:
        return datetime.strptime(value[:8], "%Y%m%d").strftime("%Y-%m-%d")
    except ValueError:
        return ""


def extract_report_header(report_nm) -> str:
    match = re.match(r"\s*\[([^\]]+)\]", report_nm or "")
    return match.group(1).strip() if match else ""


def parse_numeric_value(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(",", "").replace("%", "")
        if not normalized:
            return None
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def parse_iso_date(value):
    text = (value or "").strip()
    if not text or text == "-":
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def format_short_date(value) -> str:
    parsed = parse_iso_date(value)
    return parsed.strftime("%y-%m-%d") if parsed else ""


def normalize_text_value(value) -> str:
    if value in (None, "-"):
        return ""
    return str(value).strip()


def format_rate_text(value, prefix: str = "") -> str:
    parsed = parse_numeric_value(value)
    if parsed is None:
        return prefix.rstrip()
    return f"{prefix}{parsed:.1f}%"


def calculate_maturity_term_text(issue_date, maturity_date) -> str:
    issue_date_value = parse_iso_date(issue_date)
    maturity_date_value = parse_iso_date(maturity_date)
    if not issue_date_value or not maturity_date_value:
        return ""
    return f"{(maturity_date_value - issue_date_value).days / 365.0:.1f}년"


def clean_investor_name(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip(" ,")


def build_investor_entity_map(parsed: dict) -> dict[str, list[str]]:
    entity_map = {}
    for row in parsed.get("발행대상자세부엔티티") or []:
        if not isinstance(row, (list, tuple)) or not row:
            continue
        name = clean_investor_name(row[0])
        if not name:
            continue
        related = []
        for value in row[1:]:
            related_name = clean_investor_name(value)
            if related_name and related_name != "-" and related_name != name and related_name not in related:
                related.append(related_name)
        if related:
            entity_map[normalize_company_name_for_match(name)] = related
    return entity_map


def display_investor_name(name: str, entity_map: dict[str, list[str]]) -> str:
    if not TRUSTEE_NAME_PATTERN.search(name):
        return name
    related = entity_map.get(normalize_company_name_for_match(name)) or []
    if not related:
        return name
    return f"{name}({', '.join(related)})"


def format_investor_amount(amount) -> str:
    parsed = parse_numeric_value(amount)
    if parsed is None:
        return ""
    amount_in_eok = parsed / 10**8
    if float(amount_in_eok).is_integer():
        return f"{int(amount_in_eok):,}"
    return f"{amount_in_eok:,.1f}"


def normalize_option_schedule_triplets(schedules) -> list[list[str]]:
    normalized_rows = []
    for item in schedules or []:
        if isinstance(item, dict):
            row = [item.get("청구시작일") or "-", item.get("청구종료일") or "-", item.get("지급일") or "-"]
        elif isinstance(item, (list, tuple)):
            row = list(item[:3])
        else:
            continue
        while len(row) < 3:
            row.append("-")
        normalized_rows.append([(value or "-") for value in row[:3]])
    return normalized_rows


def get_first_option_start_date(schedules) -> str:
    triplets = normalize_option_schedule_triplets(schedules)
    if not triplets:
        return ""
    claim_start_date, _, payment_date = triplets[0]
    if claim_start_date and claim_start_date != "-":
        return claim_start_date
    return "" if payment_date == "-" else payment_date


def calculate_option_term_years(issue_date, schedules):
    triplets = normalize_option_schedule_triplets(schedules)
    if not issue_date or not triplets:
        return None
    issue_date_value = parse_iso_date(issue_date)
    payment_date_value = parse_iso_date((triplets[0][2] or "").strip())
    if not issue_date_value or not payment_date_value or payment_date_value < issue_date_value:
        return None
    return round((payment_date_value - issue_date_value).days / 365.25, 1)


def serialize_option_schedule(schedules) -> str:
    triplets = normalize_option_schedule_triplets(schedules)
    return json.dumps(triplets, ensure_ascii=False) if triplets else ""


def format_investors_text(parsed: dict) -> str:
    entity_map = build_investor_entity_map(parsed)
    issue_targets = parsed.get("발행대상자") or []
    if issue_targets:
        formatted_targets = []
        for row in issue_targets:
            if not isinstance(row, (list, tuple)) or not row:
                continue
            name = display_investor_name(clean_investor_name(row[0]), entity_map)
            amount_text = format_investor_amount(row[1] if len(row) > 1 else None)
            if not name:
                continue
            if not amount_text:
                formatted_targets.append(name)
            else:
                formatted_targets.append(f"{name} {amount_text}")
        if formatted_targets:
            return ", ".join(formatted_targets)
    investor_rows = parsed.get("투자자별투자액") or []
    if investor_rows:
        formatted = []
        for investor in investor_rows:
            name = display_investor_name(clean_investor_name(investor.get("name")), entity_map)
            amount_text = format_investor_amount(investor.get("amount"))
            if not name:
                continue
            if not amount_text:
                formatted.append(name)
            else:
                formatted.append(f"{name} {amount_text}")
        if formatted:
            return ", ".join(formatted)
    fallback = parsed.get("발행대상")
    return "" if fallback in (None, "-", "") else str(fallback).replace("\n", ", ")


def parse_report_documents(report: dict) -> dict:
    if report.get("html_path") or report.get("source_file"):
        return parse_report_document(report, [])
    return parse_report_document(report, [])


def previous_rcept_no_from_parsed(parsed: dict, current_rcept_no: str) -> str:
    families = parsed.get("correction_families") or {}
    current = str(current_rcept_no or parsed.get("rcept_no") or "").strip()
    for family in families.values():
        members = sorted(list(family.get("members") or []), key=lambda item: int(item.get("sequence") or 0))
        for index, member in enumerate(members):
            if str(member.get("rcept_no") or "").strip() == current or str(member.get("acpt_no") or "").strip() == current:
                if index == 0:
                    return ""
                return str(members[index - 1].get("rcept_no") or members[index - 1].get("acpt_no") or "")
    return ""


def build_export_row(
    report: dict,
    parsed: dict,
    company_stock_code_map: dict | None = None,
    market_cap_map: dict | None = None,
    previous_rcept_no: str = "",
) -> dict:
    company_stock_code_map = company_stock_code_map or {}
    market_cap_map = market_cap_map or {}
    issue_date = parsed.get("납입일") or ""
    security_type = infer_security_type(report.get("report_nm"), parsed.get("종류"))
    target_company_name = clean_target_stock_name(parsed.get("대상주식"))
    if not target_company_name and security_type in {"CB", "BW"}:
        target_company_name = clean_target_stock_name(report.get("corp_name"))
    target_stock_code = company_stock_code_map.get(normalize_company_name_for_match(target_company_name), "")
    issuer_stock_code = normalize_stock_code_key(report.get("stock_code"))
    market_cap_eok = market_cap_map.get(issuer_stock_code)
    put_schedules = parsed.get("PUT옵션일정표") or parsed.get("_PUT옵션일정표상세") or []
    call_schedules = parsed.get("CALL옵션일정표") or parsed.get("_CALL옵션일정표상세") or []
    filing_date = extract_filing_date_from_rcept_no(report.get("rcept_no"))
    initial_filing_date = normalize_text_value(parsed.get("최초공시일")) or extract_filing_date_from_rcept_no(previous_rcept_no) or filing_date
    maturity_date = (parsed.get("만기일") or "") if parsed.get("만기일") != "-" else ""
    refixing_floor_pct = parse_numeric_value(
        parsed.get("리픽싱(%)") if parsed.get("리픽싱(%)") is not None else parsed.get("리픽싱가격")
    )
    refixing_reason = parsed.get("리픽싱사유") or parsed.get("리픽싱내용") or ""
    refixing_reason = normalize_text_value(refixing_reason)
    premium_text = parsed.get("전환가액 결정방법") or parsed.get("할증관련텍스트") or ""
    if isinstance(premium_text, (list, tuple)):
        premium_text = " ".join(str(value) for value in premium_text if value not in (None, ""))
    premium_text = normalize_text_value(premium_text)

    return {
        "rcept_no": report.get("rcept_no", ""),
        "initial_filing_date": initial_filing_date,
        "initial_filing_date_display": format_short_date(initial_filing_date),
        "filing_date": filing_date,
        "filing_date_display": format_short_date(filing_date),
        "report_header": extract_report_header(report.get("report_nm")),
        "issuer_company_name": report.get("corp_name", ""),
        "issuer_market": ISSUER_MARKET_LABELS.get(report.get("corp_cls"), report.get("corp_cls", "")),
        "issuer_stock_code": format_issuer_stock_code(issuer_stock_code),
        "target_company_name": target_company_name,
        "target_stock_code": format_issuer_stock_code(target_stock_code),
        "round": parsed.get("회차", ""),
        "security_type": security_type,
        "issue_amount_eok": parse_numeric_value(parsed.get("발행금액") if parsed.get("발행금액") is not None else parsed.get("발행금액(억)")),
        "strike_price": parse_numeric_value(parsed.get("행사가액") if parsed.get("행사가액") is not None else parsed.get("전환가액(원)")),
        "premium_text": premium_text,
        "issue_date": issue_date if issue_date != "-" else "",
        "issue_date_display": format_short_date(issue_date if issue_date != "-" else ""),
        "maturity_date": maturity_date,
        "maturity_term_text": calculate_maturity_term_text(issue_date, maturity_date),
        "coupon_rate_pct": parse_numeric_value(parsed.get("표면이율")),
        "coupon_rate_text": format_rate_text(parsed.get("표면이율")),
        "maturity_rate_pct": parse_numeric_value(parsed.get("만기이율")),
        "maturity_rate_text": format_rate_text(parsed.get("만기이율"), prefix="/ "),
        "put_text": "/ ",
        "put_start_date": get_first_option_start_date(put_schedules),
        "put_term_years": calculate_option_term_years(issue_date, put_schedules),
        "put_ytp_pct": parse_numeric_value(parsed.get("YTP(%)")),
        "put_schedule_json": serialize_option_schedule(put_schedules),
        "call_start_date": get_first_option_start_date(call_schedules),
        "call_term_years": calculate_option_term_years(issue_date, call_schedules),
        "call_ytc_pct": parse_numeric_value(parsed.get("YTC(%)")),
        "call_cap_pct": parse_numeric_value(parsed.get("CALL행사한도(%)")),
        "call_schedule_json": serialize_option_schedule(call_schedules),
        "refixing_cycle_months": parse_numeric_value(parsed.get("리픽싱주가")),
        "refixing_price_won": parse_numeric_value(parsed.get("리픽싱(원)")),
        "refixing_floor_pct": refixing_floor_pct,
        "refixing_floor_text": format_rate_text(refixing_floor_pct),
        "refixing_reason": refixing_reason,
        "investors_text": format_investors_text(parsed),
        "venture_blank": "",
        "market_cap_eok": market_cap_eok,
        "call_blank": "",
        "sector_blank": "",
        "internal_review": "",
        "lead_blank": "",
        "underwriter": "",
        "dart_link": build_dart_link(report.get("rcept_no")),
        "disclosure_url": build_disclosure_link(report),
        "previous_rcept_no": previous_rcept_no,
    }


def _build_export_audit_item(report: dict, company_stock_code_map: dict, market_cap_map: dict | None = None) -> dict:
    parsed = {}
    parse_error = ""
    parse_failure = None
    try:
        parsed = parse_report_documents(report)
    except Exception as exc:
        parse_error = str(exc)
        parse_failure = {"rcept_no": report.get("rcept_no"), "error": parse_error}

    previous_rcept_no = ""
    family_lookup_error = ""
    family_lookup_failure = None
    try:
        previous_rcept_no = previous_rcept_no_from_parsed(parsed, report.get("rcept_no", ""))
        if not previous_rcept_no and not (report.get("html_path") or report.get("source_file")):
            previous_rcept_no = fetch_previous_family_rcept_no(report.get("rcept_no", ""))
    except Exception as exc:
        family_lookup_error = str(exc)
        family_lookup_failure = {"rcept_no": report.get("rcept_no"), "error": family_lookup_error}

    export_row = build_export_row(
        report,
        parsed,
        company_stock_code_map=company_stock_code_map,
        market_cap_map=market_cap_map,
        previous_rcept_no=previous_rcept_no,
    )
    audit_row = {
        "rcept_no": report.get("rcept_no", ""),
        "report": {
            "corp_name": report.get("corp_name", ""),
            "stock_code": report.get("stock_code", ""),
            "corp_cls": report.get("corp_cls", ""),
            "report_nm": report.get("report_nm", ""),
            "rcept_no": report.get("rcept_no", ""),
            "rcept_dt": report.get("rcept_dt", ""),
        },
        "parsed": parsed,
        "export_row": export_row,
        "error": parse_error,
        "family_lookup_error": family_lookup_error,
    }
    return {
        "report": report,
        "export_row": export_row,
        "audit_row": audit_row,
        "parse_failure": parse_failure,
        "family_lookup_failure": family_lookup_failure,
    }


def build_export_rows_with_audit(data: dict, progress_callback=None, parse_max_workers: int = DEFAULT_PARSE_WORKERS) -> tuple:
    reports = filter_reports(data.get("list", []))
    company_stock_code_map = {}
    market_cap_map = fetch_market_cap_map(progress_callback=progress_callback) if reports else {}

    rows = []
    audit_rows = []
    parse_failures = []
    family_lookup_failures = []
    total_reports = len(reports)
    parse_max_workers = max(1, int(parse_max_workers or 1))
    if reports:
        workers = min(parse_max_workers, total_reports)
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_report = {
                executor.submit(_build_export_audit_item, report, company_stock_code_map, market_cap_map): report
                for report in reports
            }
            completed = 0
            for future in as_completed(future_to_report):
                result = future.result()
                report = result["report"]
                completed += 1
                if progress_callback:
                    progress_callback(f"파싱 {completed}/{total_reports}: {report.get('corp_name', '')} {report.get('rcept_no', '')}")
                if result["parse_failure"]:
                    parse_failures.append(result["parse_failure"])
                    if progress_callback:
                        progress_callback(f"문서 파싱 실패 {report.get('rcept_no', '')}: {result['parse_failure']['error']}")
                if result["family_lookup_failure"]:
                    family_lookup_failures.append(result["family_lookup_failure"])
                    if progress_callback:
                        progress_callback(f"정정이전 조회 실패 {report.get('rcept_no', '')}: {result['family_lookup_failure']['error']}")
                rows.append(result["export_row"])
                audit_rows.append(result["audit_row"])

    summary = {
        "bgn_de": data.get("bgn_de"),
        "end_de": data.get("end_de"),
        "total_count": data.get("total_count", 0),
        "filtered_count": total_reports,
        "exported_count": len(rows),
        "parse_failure_count": len(parse_failures),
        "family_lookup_failure_count": len(family_lookup_failures),
    }
    rows.sort(key=lambda row: (row.get("filing_date") or "9999-99-99", row.get("rcept_no") or ""))
    audit_rows.sort(
        key=lambda row: (
            row.get("export_row", {}).get("filing_date") or "9999-99-99",
            row.get("export_row", {}).get("rcept_no") or "",
        )
    )
    return rows, summary, audit_rows


def export_reports(
    data: dict,
    output_path,
    audit_json_path=None,
    progress_callback=None,
    parse_max_workers: int = DEFAULT_PARSE_WORKERS,
) -> ExportResult:
    output_path = ensure_parent_dir(output_path)
    rows, summary, audit_rows = build_export_rows_with_audit(
        data,
        progress_callback=progress_callback,
        parse_max_workers=parse_max_workers,
    )
    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "reports"
    _write_summary_sheet(workbook, summary)
    _write_report_sheet(report_sheet, rows)
    workbook.save(output_path)

    audit_path = None
    if ENABLE_AUDIT_JSON:
        audit_json_path = audit_json_path or output_path.with_name(f"{output_path.stem}_audit.json")
        audit_path = save_audit_rows_json(audit_rows, audit_json_path)
    raw_path = None
    return ExportResult(output_path=output_path, audit_path=audit_path, raw_path=raw_path, summary=summary)


def save_audit_rows_json(audit_rows: list[dict], output_path) -> Path:
    output_path = ensure_parent_dir(output_path)
    output_path.write_text(json.dumps(audit_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def save_raw_reports_json(data: dict, excel_output_path) -> Path:
    excel_output_path = ensure_parent_dir(excel_output_path)
    raw_json_path = excel_output_path.with_name(f"{excel_output_path.stem}_raw.json")
    raw_json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return raw_json_path


def _write_summary_sheet(workbook, summary: dict) -> None:
    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.append(["key", "value"])
    for key, value in summary.items():
        summary_sheet.append([key, value])
    summary_sheet.freeze_panes = "A2"


def _write_report_sheet(sheet, rows: list[dict]) -> None:
    sheet.page_margins.left = 0.1
    sheet.page_margins.right = 0.1
    _write_header(sheet)
    _write_rows(sheet, rows)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def _write_header(sheet) -> None:
    for index, spec in enumerate(COLUMN_SPECS, start=1):
        _style_header_cell(sheet.cell(row=1, column=index, value=spec["header"]))
        sheet.column_dimensions[get_column_letter(index)].width = spec["width"]
    sheet.row_dimensions[1].height = 22


def _style_header_cell(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGNMENT
    cell.border = HEADER_BORDER


def _write_rows(sheet, rows: list[dict]) -> None:
    for row_index, row in enumerate(rows, start=2):
        for column_index, spec in enumerate(COLUMN_SPECS, start=1):
            value = _display_cell_value(row.get(spec["key"], ""))
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            _style_body_cell(cell, spec)
        sheet.row_dimensions[row_index].height = 42


def _display_cell_value(value):
    if value is None:
        return BLANK_CELL_VALUE
    if isinstance(value, str) and not value.strip():
        return BLANK_CELL_VALUE
    return value


def _style_body_cell(cell, spec: dict) -> None:
    key = spec["key"]
    cell.alignment = Alignment(
        horizontal=spec.get("align", "left"),
        vertical="center",
        wrap_text=True,
        indent=1 if spec.get("margin") else 0,
    )
    cell.border = BODY_BORDER
    if key in INTEGER_FORMAT_KEYS and isinstance(cell.value, (int, float)):
        cell.number_format = "#,##0"
    elif key in DECIMAL_FORMAT_KEYS and isinstance(cell.value, (int, float)):
        cell.number_format = "0.0"
    elif key == "disclosure_url" and cell.value != BLANK_CELL_VALUE:
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

    if cell.style == "Hyperlink":
        cell.font = Font(size=9, color="0563C1", underline="single")
    else:
        cell.font = BODY_FONT
