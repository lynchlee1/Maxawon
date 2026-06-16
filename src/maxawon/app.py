from __future__ import annotations

import csv
import os
import platform
import shutil
import signal
import subprocess
import threading
from pathlib import Path
from tkinter import DISABLED, END, NORMAL, StringVar, Tk, filedialog, messagebox, ttk

from maxawon.scrapling_adapter import check_scrapling
from maxawon.table_capture import (
    CDP_URL,
    CaptureResult,
    CapturedTable,
    capture_current_maxawon_table_sync,
    write_table_csv,
)


MAXAWON_URL = "https://www.maxawon.com/"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
PROFILE_DIR = PROJECT_ROOT / ".chrome-profile"
DEFAULT_CAPTURE_OUTPUT = PROJECT_ROOT / "output" / "maxawon_condition_search.csv"
REMOTE_DEBUGGING_PORT = "9222"

COLOR_BG = "#eef2f7"
COLOR_SURFACE = "#ffffff"
COLOR_SURFACE_MUTED = "#f8fafc"
COLOR_BORDER = "#d7dde8"
COLOR_TEXT = "#172033"
COLOR_MUTED = "#667085"
COLOR_ACCENT = "#2563eb"
COLOR_ACCENT_ACTIVE = "#1d4ed8"
COLOR_DANGER = "#b42318"
COLOR_DANGER_ACTIVE = "#8f1d15"
COLOR_SIDEBAR = "#111827"
COLOR_SIDEBAR_ACTIVE = "#1f2937"
COLOR_SIDEBAR_TEXT = "#e5e7eb"


def find_chrome() -> str | None:
    system = platform.system()
    if system == "Darwin":
        candidates = [
            Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
            Path.home() / "Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        ]
    elif system == "Windows":
        candidates = [
            Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
            Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
            Path.home() / r"AppData\Local\Google\Chrome\Application\chrome.exe",
        ]
    else:
        for command in ("google-chrome", "google-chrome-stable", "chromium", "chromium-browser"):
            chrome = shutil.which(command)
            if chrome is not None:
                return chrome
        candidates = []

    for candidate in candidates:
        if candidate.exists():
            return str(candidate)

    return None


def _matching_pids(patterns: list[str]) -> list[int]:
    pids: set[int] = set()
    for pattern in patterns:
        result = subprocess.run(["pgrep", "-f", pattern], capture_output=True, text=True, check=False)
        if result.returncode not in (0, 1):
            continue
        for line in result.stdout.splitlines():
            try:
                pid = int(line.strip())
            except ValueError:
                continue
            if pid > 0 and pid != os.getpid():
                pids.add(pid)
    return sorted(pids)


def close_app_chrome_processes() -> int:
    patterns = [str(PROFILE_DIR), f"--remote-debugging-port={REMOTE_DEBUGGING_PORT}"]
    if platform.system() == "Windows":
        profile = str(PROFILE_DIR).replace("'", "''")
        port = f"--remote-debugging-port={REMOTE_DEBUGGING_PORT}"
        script = f"""
$matches = Get-CimInstance Win32_Process | Where-Object {{
  $_.Name -eq 'chrome.exe' -and $_.CommandLine -and (
    $_.CommandLine -like '*{profile}*' -or $_.CommandLine -like '*{port}*'
  )
}}
$matches | ForEach-Object {{
  Stop-Process -Id $_.ProcessId -Force
  $_.ProcessId
}}
"""
        result = subprocess.run(
            ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or "앱이 연 Chrome 프로세스 종료에 실패했습니다.")
        return len([line for line in result.stdout.splitlines() if line.strip()])

    count = 0
    for pid in _matching_pids(patterns):
        try:
            os.kill(pid, signal.SIGTERM)
            count += 1
        except ProcessLookupError:
            pass
    return count


def close_all_chrome_processes() -> int:
    system = platform.system()
    if system == "Darwin":
        result = subprocess.run(["pkill", "-f", "Google Chrome"], capture_output=True, text=True, check=False)
    elif system == "Windows":
        script = """
$matches = Get-CimInstance Win32_Process | Where-Object { $_.Name -eq 'chrome.exe' }
$matches | ForEach-Object {
  Stop-Process -Id $_.ProcessId -Force
  $_.ProcessId
}
"""
        result = subprocess.run(
            ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or "Chrome 전체 종료에 실패했습니다.")
        return len([line for line in result.stdout.splitlines() if line.strip()])
    else:
        result = subprocess.run(["pkill", "-f", "chrome|chromium"], capture_output=True, text=True, check=False)

    if result.returncode not in (0, 1):
        raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "Chrome 전체 종료에 실패했습니다.")
    return 1 if result.returncode == 0 else 0


def read_excel_preview(path: Path, limit: int = 20) -> tuple[list[str], list[list[str]]]:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        last_error: UnicodeDecodeError | None = None
        for encoding in ("utf-8-sig", "cp949"):
            try:
                with path.open("r", encoding=encoding, newline="") as file:
                    reader = csv.reader(file)
                    rows = list(reader)
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            raise RuntimeError("CSV 파일 인코딩을 읽지 못했습니다.") from last_error
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "엑셀(.xlsx) 파일을 읽으려면 openpyxl이 필요합니다. "
                "터미널에서 `python -m pip install -e .`를 실행하세요."
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [
            ["" if value is None else str(value) for value in row]
            for row in sheet.iter_rows(max_row=limit + 1, values_only=True)
        ]
        workbook.close()

    if not rows:
        return [], []

    headers = rows[0]
    body = rows[1 : limit + 1]
    return headers, body


class MaxawonApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.root.title("Maxawon")
        self.root.geometry("1080x720")
        self.root.minsize(900, 620)

        self.excel_path: Path | None = None
        self.capture_output_path = DEFAULT_CAPTURE_OUTPUT
        self.login_status = StringVar(value="로그인 전")
        self.file_status = StringVar(value="엑셀 파일 미선택")
        self.progress_status = StringVar(value="대기 중")
        self.scrapling_status = StringVar(value="확인 전")
        self.capture_status = StringVar(value="대기 중")
        self.capture_output_status = StringVar(value=str(self.capture_output_path))
        self.capture_max_pages = StringVar(value="30")
        self.nav_buttons: dict[str, ttk.Button] = {}
        self.views: dict[str, ttk.Frame] = {}

        self._build()

    def _build(self) -> None:
        self._configure_style()

        self.root.configure(bg=COLOR_BG)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        top = ttk.Frame(self.root, padding=(24, 20, 24, 16), style="App.TFrame")
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        ttk.Label(top, text="Maxawon", style="Title.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(top, text="수동 로그인 기반의 Maxawon 업무 자동화 콘솔", style="Muted.TLabel").grid(
            row=1, column=0, sticky="w", pady=(6, 0)
        )

        body = ttk.Frame(self.root, padding=(24, 0), style="App.TFrame")
        body.grid(row=1, column=0, sticky="nsew")
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        sidebar = ttk.Frame(body, padding=(14, 16), style="Sidebar.TFrame")
        sidebar.grid(row=0, column=0, sticky="ns", padx=(0, 16))

        content = ttk.Frame(body, style="App.TFrame")
        content.grid(row=0, column=1, sticky="nsew")
        content.columnconfigure(0, weight=1)
        content.rowconfigure(0, weight=1)

        session_tab = ttk.Frame(content, padding=0, style="App.TFrame")
        capture_tab = ttk.Frame(content, padding=0, style="App.TFrame")
        excel_tab = ttk.Frame(content, padding=0, style="App.TFrame")
        for name, frame in (
            ("session", session_tab),
            ("capture", capture_tab),
            ("excel", excel_tab),
        ):
            frame.grid(row=0, column=0, sticky="nsew")
            self.views[name] = frame

        self._build_sidebar(sidebar)
        self._build_session_tab(session_tab)
        self._build_capture_tab(capture_tab)
        self._build_excel_tab(excel_tab)
        self._show_view("session")

        log_frame = ttk.Frame(self.root, padding=16, style="Card.TFrame")
        log_frame.grid(row=2, column=0, sticky="ew", padx=24, pady=(16, 24))
        log_frame.columnconfigure(0, weight=1)

        ttk.Label(log_frame, text="활동 로그", style="CardTitle.TLabel").grid(
            row=0, column=0, sticky="w", pady=(0, 10)
        )
        self.log = ttk.Treeview(log_frame, columns=("message",), show="headings", height=5)
        self.log.heading("message", text="메시지")
        self.log.column("message", width=900, stretch=True)
        self.log.grid(row=1, column=0, sticky="ew")

        self._set_excel_start_enabled()
        self.add_log("Chrome을 열고 직접 로그인한 뒤 '로그인 완료'를 누르세요.")
        self.check_scrapling_status(show_popup=False)

    def _configure_style(self) -> None:
        style = ttk.Style(self.root)
        if "clam" in style.theme_names():
            style.theme_use("clam")

        style.configure(".", font=("", 13), background=COLOR_BG, foreground=COLOR_TEXT)
        style.configure("App.TFrame", background=COLOR_BG)
        style.configure("Card.TFrame", background=COLOR_SURFACE, borderwidth=1, relief="solid")
        style.configure("Sidebar.TFrame", background=COLOR_SIDEBAR)
        style.configure("TLabel", background=COLOR_BG, foreground=COLOR_TEXT)
        style.configure("Title.TLabel", background=COLOR_BG, foreground=COLOR_TEXT, font=("", 24, "bold"))
        style.configure("Muted.TLabel", background=COLOR_BG, foreground=COLOR_MUTED, font=("", 13))
        style.configure("CardTitle.TLabel", background=COLOR_SURFACE, foreground=COLOR_TEXT, font=("", 15, "bold"))
        style.configure("CardMuted.TLabel", background=COLOR_SURFACE, foreground=COLOR_MUTED, font=("", 12))
        style.configure("Field.TLabel", background=COLOR_SURFACE, foreground=COLOR_MUTED, font=("", 12))
        style.configure("Value.TLabel", background=COLOR_SURFACE, foreground=COLOR_TEXT, font=("", 13, "bold"))
        style.configure("SidebarTitle.TLabel", background=COLOR_SIDEBAR, foreground=COLOR_SIDEBAR_TEXT, font=("", 13, "bold"))
        style.configure("SidebarMuted.TLabel", background=COLOR_SIDEBAR, foreground="#9ca3af", font=("", 11))

        style.configure("TButton", padding=(14, 8), borderwidth=0, focusthickness=0)
        style.configure("Accent.TButton", background=COLOR_ACCENT, foreground="#ffffff")
        style.map("Accent.TButton", background=[("active", COLOR_ACCENT_ACTIVE), ("disabled", "#9ca3af")])
        style.configure("Danger.TButton", background=COLOR_DANGER, foreground="#ffffff")
        style.map("Danger.TButton", background=[("active", COLOR_DANGER_ACTIVE), ("disabled", "#9ca3af")])
        style.configure("Secondary.TButton", background=COLOR_SURFACE_MUTED, foreground=COLOR_TEXT)
        style.map("Secondary.TButton", background=[("active", "#e5e7eb")])
        style.configure("Nav.TButton", anchor="w", padding=(12, 10), background=COLOR_SIDEBAR, foreground=COLOR_SIDEBAR_TEXT)
        style.configure("NavActive.TButton", anchor="w", padding=(12, 10), background=COLOR_SIDEBAR_ACTIVE, foreground="#ffffff")
        style.map("Nav.TButton", background=[("active", COLOR_SIDEBAR_ACTIVE)])
        style.map("NavActive.TButton", background=[("active", COLOR_SIDEBAR_ACTIVE)])

        style.configure("TSpinbox", fieldbackground=COLOR_SURFACE, bordercolor=COLOR_BORDER, lightcolor=COLOR_BORDER, darkcolor=COLOR_BORDER)
        style.configure(
            "Treeview",
            background=COLOR_SURFACE,
            fieldbackground=COLOR_SURFACE,
            foreground=COLOR_TEXT,
            bordercolor=COLOR_BORDER,
            rowheight=30,
        )
        style.configure("Treeview.Heading", background=COLOR_SURFACE_MUTED, foreground=COLOR_MUTED, font=("", 12, "bold"))
        style.map("Treeview", background=[("selected", "#dbeafe")], foreground=[("selected", COLOR_TEXT)])

    def _build_sidebar(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(0, weight=1)

        ttk.Label(parent, text="WORKSPACE", style="SidebarMuted.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(parent, text="Maxawon", style="SidebarTitle.TLabel").grid(row=1, column=0, sticky="w", pady=(4, 18))

        nav_items = [
            ("session", "세션 연결"),
            ("capture", "조건검색 복사"),
            ("excel", "엑셀 처리"),
        ]
        for row, (name, label) in enumerate(nav_items, start=2):
            button = ttk.Button(
                parent,
                text=label,
                style="Nav.TButton",
                command=lambda view=name: self._show_view(view),
                width=18,
            )
            button.grid(row=row, column=0, sticky="ew", pady=(0, 8))
            self.nav_buttons[name] = button

    def _show_view(self, name: str) -> None:
        self.views[name].tkraise()
        for view_name, button in self.nav_buttons.items():
            button.configure(style="NavActive.TButton" if view_name == name else "Nav.TButton")

    def _make_card(self, parent: ttk.Frame, title: str, description: str | None = None) -> ttk.Frame:
        card = ttk.Frame(parent, padding=18, style="Card.TFrame")
        card.columnconfigure(0, weight=1)
        ttk.Label(card, text=title, style="CardTitle.TLabel").grid(row=0, column=0, sticky="w")
        if description is not None:
            ttk.Label(card, text=description, style="CardMuted.TLabel").grid(
                row=1, column=0, sticky="w", pady=(4, 14)
            )
        return card

    def _build_session_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(0, weight=1)

        actions = self._make_card(
            parent,
            "세션 연결",
            "Chrome을 원격 디버깅 모드로 열고, 로그인은 사용자가 직접 완료합니다.",
        )
        actions.grid(row=0, column=0, sticky="ew")
        actions.columnconfigure(4, weight=1)

        ttk.Button(actions, text="Chrome 열기", style="Accent.TButton", command=self.open_chrome).grid(
            row=2, column=0, padx=(0, 8), sticky="w"
        )
        ttk.Button(actions, text="로그인 완료", style="Secondary.TButton", command=self.mark_login_done).grid(
            row=2, column=1, padx=(0, 8), sticky="w"
        )
        ttk.Button(actions, text="실행된 Chrome 종료", style="Danger.TButton", command=self.close_app_chrome).grid(
            row=2, column=2, padx=(0, 8), sticky="w"
        )
        ttk.Button(actions, text="전체 Chrome 종료", style="Danger.TButton", command=self.close_all_chrome).grid(
            row=2, column=3, padx=(0, 8), sticky="w"
        )
        ttk.Button(actions, text="Scrapling 확인", style="Secondary.TButton", command=self.check_scrapling_status).grid(
            row=2, column=4, padx=(0, 8), sticky="w"
        )

        status = self._make_card(parent, "상태")
        status.grid(row=1, column=0, sticky="ew", pady=(16, 0))
        status.columnconfigure(1, weight=1)

        self._status_row(status, 1, "로그인 상태", self.login_status)
        self._status_row(status, 2, "진행 상태", self.progress_status)
        self._status_row(status, 3, "Scrapling", self.scrapling_status)
        self._status_row(status, 4, "브라우저 연결", CDP_URL)

    def _build_capture_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)

        controls = self._make_card(
            parent,
            "현재 조건검색 결과",
            "Maxawon에서 사용자가 직접 띄운 조건검색 결과 테이블을 CSV로 저장합니다.",
        )
        controls.grid(row=0, column=0, sticky="ew")
        controls.columnconfigure(1, weight=1)

        ttk.Label(controls, text="최대 페이지", style="Field.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 8))
        ttk.Spinbox(
            controls,
            from_=1,
            to=500,
            textvariable=self.capture_max_pages,
            width=8,
        ).grid(row=2, column=1, sticky="w")
        self.capture_button = ttk.Button(
            controls,
            text="화면 테이블 복사",
            style="Accent.TButton",
            command=self.start_table_capture,
        )
        self.capture_button.grid(row=2, column=2, sticky="e")

        ttk.Label(controls, text="저장 파일", style="Field.TLabel").grid(row=3, column=0, sticky="w", padx=(0, 8), pady=(12, 0))
        ttk.Label(controls, textvariable=self.capture_output_status, style="Value.TLabel").grid(
            row=3, column=1, sticky="ew", pady=(12, 0)
        )
        ttk.Button(controls, text="변경", style="Secondary.TButton", command=self.pick_capture_output).grid(
            row=3, column=2, sticky="e", pady=(12, 0)
        )

        status = ttk.Frame(parent, style="App.TFrame")
        status.grid(row=1, column=0, sticky="ew", pady=(12, 8))
        status.columnconfigure(1, weight=1)
        ttk.Label(status, text="복사 상태", style="Muted.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(status, textvariable=self.capture_status, style="Muted.TLabel").grid(row=0, column=1, sticky="w")

        preview_frame = self._make_card(parent, "복사 결과 미리보기")
        preview_frame.grid(row=2, column=0, sticky="nsew")
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(1, weight=1)

        self.capture_preview = ttk.Treeview(preview_frame, show="headings", height=14)
        self.capture_preview.grid(row=1, column=0, sticky="nsew", pady=(12, 0))
        self._attach_tree_scrollbars(preview_frame, self.capture_preview, row=1)

    def _build_excel_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)

        actions = self._make_card(
            parent,
            "엑셀 처리",
            "검색 대상 파일을 불러오고 앞부분을 확인합니다. 자동 검색 처리는 규칙 확정 후 진행됩니다.",
        )
        actions.grid(row=0, column=0, sticky="ew")
        actions.columnconfigure(1, weight=1)

        ttk.Button(actions, text="엑셀 선택", style="Accent.TButton", command=self.pick_excel).grid(
            row=2, column=0, padx=(0, 8), sticky="w"
        )
        self.start_button = ttk.Button(actions, text="처리 시작", style="Secondary.TButton", command=self.start_processing)
        self.start_button.grid(row=2, column=2, sticky="e")

        status_grid = ttk.Frame(parent, style="App.TFrame")
        status_grid.grid(row=1, column=0, sticky="ew", pady=(12, 8))
        status_grid.columnconfigure(1, weight=1)

        ttk.Label(status_grid, text="파일 상태", style="Muted.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10))
        ttk.Label(status_grid, textvariable=self.file_status, style="Muted.TLabel").grid(row=0, column=1, sticky="w")
        ttk.Label(status_grid, text="진행 상태", style="Muted.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(4, 0))
        ttk.Label(status_grid, textvariable=self.progress_status, style="Muted.TLabel").grid(row=1, column=1, sticky="w", pady=(4, 0))

        preview_frame = self._make_card(parent, "엑셀 미리보기")
        preview_frame.grid(row=2, column=0, sticky="nsew")
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(1, weight=1)

        self.preview = ttk.Treeview(preview_frame, show="headings", height=12)
        self.preview.grid(row=1, column=0, sticky="nsew", pady=(12, 0))
        self._attach_tree_scrollbars(preview_frame, self.preview, row=1)

    def _status_row(self, parent: ttk.Frame, row: int, label: str, value: StringVar | str) -> None:
        ttk.Label(parent, text=label, style="Field.TLabel").grid(row=row, column=0, sticky="w", padx=(0, 18), pady=(10, 0))
        ttk.Label(parent, textvariable=value if isinstance(value, StringVar) else None, text=None if isinstance(value, StringVar) else value, style="Value.TLabel").grid(
            row=row, column=1, sticky="w", pady=(10, 0)
        )

    def _attach_tree_scrollbars(self, parent: ttk.Frame, tree: ttk.Treeview, row: int = 0) -> None:
        y_scroll = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        y_scroll.grid(row=row, column=1, sticky="ns", pady=(12 if row else 0, 0))
        x_scroll = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        x_scroll.grid(row=row + 1, column=0, sticky="ew")
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

    def open_chrome(self) -> None:
        chrome = find_chrome()
        if chrome is None:
            messagebox.showerror("Chrome 없음", "이 PC에서 Chrome 실행 파일을 찾지 못했습니다.")
            return

        PROFILE_DIR.mkdir(exist_ok=True)
        subprocess.Popen(
            [
                chrome,
                f"--user-data-dir={PROFILE_DIR}",
                f"--remote-debugging-port={REMOTE_DEBUGGING_PORT}",
                "--new-window",
                MAXAWON_URL,
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        self.progress_status.set("Chrome 실행됨")
        self.add_log("Chrome을 열었습니다. Maxawon에서 직접 로그인한 뒤 조건검색을 실행하세요.")

    def close_app_chrome(self) -> None:
        try:
            count = close_app_chrome_processes()
        except RuntimeError as exc:
            messagebox.showerror("Chrome 종료 실패", str(exc))
            self.add_log(str(exc))
            return

        self.login_status.set("로그인 전")
        self.progress_status.set("실행된 Chrome 종료됨")
        self.add_log(
            "앱이 연 Chrome 프로세스를 종료했습니다."
            if count
            else "종료할 앱 Chrome 프로세스를 찾지 못했습니다."
        )
        self._set_excel_start_enabled()

    def close_all_chrome(self) -> None:
        if not messagebox.askyesno("전체 Chrome 종료", "사용자가 직접 연 Chrome까지 모두 종료합니다. 계속할까요?"):
            return

        try:
            count = close_all_chrome_processes()
        except RuntimeError as exc:
            messagebox.showerror("Chrome 종료 실패", str(exc))
            self.add_log(str(exc))
            return

        self.login_status.set("로그인 전")
        self.progress_status.set("전체 Chrome 종료됨")
        self.add_log(
            "모든 Chrome 프로세스 종료 명령을 실행했습니다."
            if count
            else "실행 중인 Chrome 프로세스를 찾지 못했습니다."
        )
        self._set_excel_start_enabled()

    def mark_login_done(self) -> None:
        self.login_status.set("로그인 완료")
        self.progress_status.set("엑셀 파일 선택 대기")
        self.add_log("사용자가 로그인 완료를 확인했습니다.")
        self._set_excel_start_enabled()

    def check_scrapling_status(self, show_popup: bool = True) -> None:
        status = check_scrapling()
        self.scrapling_status.set(status.message)
        self.add_log(status.message)

        if show_popup:
            if status.installed:
                messagebox.showinfo("Scrapling 확인", status.message)
            else:
                messagebox.showwarning("Scrapling 필요", status.message)

    def pick_excel(self) -> None:
        selected = filedialog.askopenfilename(
            title="검색 대상 파일 선택",
            filetypes=[
                ("Excel or CSV", "*.xlsx *.xlsm *.csv"),
                ("Excel", "*.xlsx *.xlsm"),
                ("CSV", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if not selected:
            return

        path = Path(selected)
        try:
            headers, rows = read_excel_preview(path)
        except Exception as exc:
            messagebox.showerror("파일 읽기 실패", str(exc))
            return

        self.excel_path = path
        self.file_status.set(str(path))
        self.progress_status.set("엑셀 파일 로드됨")
        self.render_preview(headers, rows)
        self.add_log(f"검색 대상 파일을 불러왔습니다: {path.name}")
        self._set_excel_start_enabled()

    def render_preview(self, headers: list[str], rows: list[list[str]]) -> None:
        self.render_table(self.preview, headers, rows)

    def render_table(self, tree: ttk.Treeview, headers: list[str], rows: list[list[str]]) -> None:
        tree.delete(*tree.get_children())
        column_ids = [f"col_{index}" for index in range(len(headers))]
        tree["columns"] = column_ids

        for index, column_id in enumerate(column_ids):
            header = headers[index] or f"Column {index + 1}"
            tree.heading(column_id, text=header)
            tree.column(column_id, width=140, minwidth=80, stretch=True)

        for row in rows:
            values = row + [""] * (len(headers) - len(row))
            tree.insert("", END, values=values[: len(headers)])

    def pick_capture_output(self) -> None:
        selected = filedialog.asksaveasfilename(
            title="조건검색 복사 결과 저장",
            defaultextension=".csv",
            filetypes=[
                ("CSV", "*.csv"),
                ("All files", "*.*"),
            ],
            initialfile=self.capture_output_path.name,
        )
        if not selected:
            return

        self.capture_output_path = Path(selected)
        self.capture_output_status.set(str(self.capture_output_path))

    def start_table_capture(self) -> None:
        if self.login_status.get() != "로그인 완료":
            messagebox.showwarning("로그인 필요", "Maxawon에 로그인한 뒤 '로그인 완료'를 누르세요.")
            return

        try:
            max_pages = int(self.capture_max_pages.get())
        except ValueError:
            messagebox.showwarning("페이지 수 확인", "최대 페이지는 숫자로 입력하세요.")
            return
        if max_pages < 1:
            messagebox.showwarning("페이지 수 확인", "최대 페이지는 1 이상이어야 합니다.")
            return

        self.capture_button.configure(state=DISABLED)
        self.capture_status.set("복사 중")
        self.add_log("현재 Maxawon 화면의 조건검색 결과 테이블 복사를 시작합니다.")
        output_path = self.capture_output_path

        thread = threading.Thread(
            target=self._run_table_capture,
            args=(max_pages, output_path),
            daemon=True,
        )
        thread.start()

    def _run_table_capture(self, max_pages: int, output_path: Path) -> None:
        try:
            result = capture_current_maxawon_table_sync(max_pages=max_pages)
            if not result.rows:
                raise RuntimeError("현재 화면에서 복사할 테이블 데이터를 찾지 못했습니다.")
            output_path.parent.mkdir(parents=True, exist_ok=True)
            write_table_csv(
                output_path,
                CapturedTable(result.headers, result.rows),
            )
        except Exception as exc:
            self.root.after(0, self._finish_table_capture_error, str(exc))
            return

        self.root.after(0, self._finish_table_capture_success, result, output_path)

    def _finish_table_capture_success(self, result: CaptureResult, output_path: Path) -> None:
        self.render_table(self.capture_preview, result.headers, result.rows[:100])
        self.capture_status.set(
            f"{result.pages}페이지, {len(result.rows)}행 저장 완료"
        )
        self.capture_button.configure(state=NORMAL)
        self.add_log(f"조건검색 결과를 저장했습니다: {output_path}")

    def _finish_table_capture_error(self, message: str) -> None:
        self.capture_status.set("복사 실패")
        self.capture_button.configure(state=NORMAL)
        self.add_log(f"조건검색 결과 복사 실패: {message}")
        messagebox.showerror("복사 실패", message)

    def start_processing(self) -> None:
        if self.login_status.get() != "로그인 완료":
            messagebox.showwarning("로그인 필요", "먼저 Maxawon에 로그인한 뒤 '로그인 완료'를 누르세요.")
            return
        if self.excel_path is None:
            messagebox.showwarning("파일 필요", "검색 대상 엑셀 파일을 선택하세요.")
            return
        scrapling = check_scrapling()
        if not scrapling.installed:
            self.scrapling_status.set(scrapling.message)
            messagebox.showwarning("Scrapling 필요", scrapling.message)
            return

        self.progress_status.set("구현 대기")
        self.add_log("Scrapling 기반 검색 처리는 아직 구현되지 않았습니다. 중복 후보 처리 규칙을 먼저 확정해야 합니다.")
        messagebox.showinfo(
            "아직 구현 전",
            "Scrapling 기반 엑셀 검색 처리는 아직 보류 상태입니다.\n"
            "기업명/법인번호 컬럼, 출력 항목, 중복 후보 선택 기준을 확정한 뒤 구현하세요.",
        )

    def add_log(self, message: str) -> None:
        self.log.insert("", END, values=(message,))
        children = self.log.get_children()
        if children:
            self.log.see(children[-1])

    def _set_excel_start_enabled(self) -> None:
        if self.login_status.get() == "로그인 완료" and self.excel_path is not None:
            self.start_button.configure(state=NORMAL)
        else:
            self.start_button.configure(state=DISABLED)


def main() -> int:
    root = Tk()
    MaxawonApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
